import pandas as pd
import os

file_path = "sii_data/generados/Planilla_test_job_manual_047.xlsx"

if os.path.exists(file_path):
    print(f"Inspecting {file_path}...")
    try:
        xls = pd.ExcelFile(file_path)
        sheets = xls.sheet_names
        print(f"Sheets found: {sheets}")
        
        target_sheets = ['RCV Venta', 'RCV Compra']

        for sheet in target_sheets:
            if sheet in sheets:
                print(f"\n--- Sheet: {sheet} ---")
                df = pd.read_excel(xls, sheet_name=sheet, header=None)
                
                # Check columns of the second table
                if sheet in ['RCV Venta', 'RCV Compra']:
                    # Assuming header is at row 12 (index 11) or so
                    # Let's find the row starting with 'Nro'
                    header_row_idx = None
                    for idx, row in df.iterrows():
                        if str(row[0]).strip() == 'Nro':
                            header_row_idx = idx
                            break
                    
                    if header_row_idx is not None:
                        print(f"\nFound Detail Header at row {header_row_idx}")
                        headers = df.iloc[header_row_idx].tolist()
                        print(f"Columns found: {headers}")
                        
                        # Print first few rows of data
                        print("\nFirst 3 rows of data:")
                        data_start = header_row_idx + 1
                        if data_start < len(df):
                            print(df.iloc[data_start:data_start+3])
                            
                        # Check specific columns
                        try:
                            if sheet == 'RCV Venta':
                                # Map column name to index
                                if 'Codigo Sucursal' in headers:
                                    col_idx = headers.index('Codigo Sucursal')
                                    print(f"\nCodigo Sucursal values (first 5):")
                                    print(df.iloc[data_start:data_start+5, col_idx])
                                
                                if 'Tipo Doc' in headers:
                                    doc_idx = headers.index('Tipo Doc')
                                    print(f"\nTipo Doc values (first 5):")
                                    print(df.iloc[data_start:data_start+5, doc_idx])
                                
                                if 'Tipo Venta' in headers:
                                    tv_idx = headers.index('Tipo Venta')
                                    print(f"\nTipo Venta values (first 5):")
                                    print(df.iloc[data_start:data_start+5, tv_idx])

                            elif sheet == 'RCV Compra':
                                # Check RCV Compra specific columns
                                target_cols_compra = ['Tipo Doc', 'Tipo Compra', 'Monto IVA Recuperable']
                                for col in target_cols_compra:
                                    if col in headers:
                                        c_idx = headers.index(col)
                                        print(f"\n{col} values (first 5):")
                                        print(df.iloc[data_start:data_start+5, c_idx])
                                    else:
                                        print(f"\nColumn {col} NOT FOUND in RCV Compra")

                        except ValueError:
                            print("Column not found in headers")
                    else:
                        print("\nCould not find 'Nro' header row")
            else:
                print(f"\n--- Sheet: {sheet} NOT FOUND ---")
        
        # Check Images
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        
        for sname in ['RCV Compra', 'Compras', 'Ventas', 'BH', 'IMPUESTO UNICO', 'DIN']:
            if sname in wb.sheetnames:
                ws = wb[sname]
                print(f"\n--- {sname} Images ---")
                print(f"Images count: {len(ws._images)}")
        
        # Check DIN Content
        if 'DIN' in sheets:
            print("\n--- Sheet: DIN ---")
            df = pd.read_excel(xls, sheet_name='DIN', header=None)
            print(f"DIN Sheet Shape: {df.shape}")
        
        # Check Compras Content
        if 'Compras' in sheets:
            print("\n--- Sheet: Compras (Dashboard) ---")
            df = pd.read_excel(xls, sheet_name='Compras', header=None)
            print(f"Cell B2 (Title): {df.iloc[1, 1]}")
            
        # Check Ventas Content
        if 'Ventas' in sheets:
            print("\n--- Sheet: Ventas (Dashboard) ---")
            df = pd.read_excel(xls, sheet_name='Ventas', header=None)
            print(f"Cell B2 (Title): {df.iloc[1, 1]}")
            print(f"Cell B18 (Table 1 Header): {df.iloc[17, 1]}")
            print(f"Cell G18 (Table 2 Header): {df.iloc[17, 6]}")
            print(f"Cell B28 (MES VS AÑO Label): {df.iloc[27, 1]}")
            print(f"Cell E28 (Percentage): {df.iloc[27, 4]}")
            
        # Check BH Content
        if 'BH' in sheets:
            print("\n--- Sheet: BH ---")
            df = pd.read_excel(xls, sheet_name='BH', header=None)
            print(f"Cell I1 (Brutos): {df.iloc[0, 8]}")
            print(f"Cell F2 (Label Totales): {df.iloc[1, 5]}")
            print(f"Cell A4 (Contribuyente): {df.iloc[3, 0]}")
            print(f"Cell A8 (Header Boleta): {df.iloc[7, 0]}")
            print(f"Cell A9 (Subheader N): {df.iloc[8, 0]}")
             
        # Check IMPUESTO UNICO Content
        if 'IMPUESTO UNICO' in sheets:
            print("\n--- Sheet: IMPUESTO UNICO ---")
            df = pd.read_excel(xls, sheet_name='IMPUESTO UNICO', header=None)
            print(f"Cell A8 (Header): {df.iloc[7, 0]}")
            print(f"Cell C8 (Header): {df.iloc[7, 2]}")

         # Check Analisis Sheets
        analisis_sheets = ['Resumen', 'Compras analisis', 'Ventas Analisis', 'Boletas']
        for sheet in analisis_sheets:
            if sheet in sheets:
                print(f"\n--- Sheet: {sheet} ---")
                if sheet == 'Compras analisis' or sheet == 'Ventas Analisis':
                    df = pd.read_excel(xls, sheet_name=sheet, header=None)
                    print(f"Sheet Shape: {df.shape}")
                    print(f"Cell A1: {df.iloc[0, 0]}")
                    print(f"Cell E1: {df.iloc[0, 4]}")
                    print(f"Cell H1: {df.iloc[0, 7]}")
                elif sheet == 'Boletas' or sheet == 'Resumen':
                     # Read first few rows
                     df = pd.read_excel(xls, sheet_name=sheet, header=None)
                     print(f"{sheet} Raw Head (First 20 rows):\n{df.head(20)}")
                else:
                    df = pd.read_excel(xls, sheet_name=sheet)
                    print(f"Columns: {df.columns.tolist()}")
                    print("First 3 rows:")
                    print(df.head(3))
            else:
                print(f"\n--- Sheet: {sheet} NOT FOUND ---")
                
    except Exception as e:
        print(f"Error: {e}")
else:
    print(f"File not found: {file_path}")
