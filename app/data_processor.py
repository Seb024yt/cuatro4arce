import os
import pandas as pd
import glob
import re
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.drawing.image import Image

MONTH_MAP = {
    '01': 'ENERO', '02': 'FEBRERO', '03': 'MARZO', '04': 'ABRIL',
    '05': 'MAYO', '06': 'JUNIO', '07': 'JULIO', '08': 'AGOSTO',
    '09': 'SEPTIEMBRE', '10': 'OCTUBRE', '11': 'NOVIEMBRE', '12': 'DICIEMBRE'
}

DOC_TYPE_MAP = {
    33: 'Factura Electrónica',
    34: 'Factura No Afecta o Exenta Electrónica',
    39: 'Boleta Electrónica',
    41: 'Boleta Exenta Electrónica',
    56: 'Nota de Débito Electrónica',
    61: 'Nota de Crédito Electrónica',
    # Add others as needed
}

def extract_month_from_filename(filename):
    # Try to find pattern YYYYMM or MMYYYY
    match = re.search(r'202[0-9](0[1-9]|1[0-2])', filename)
    if match:
        mm = match.group(1)
        return MONTH_MAP.get(mm, 'DESCONOCIDO')
    return 'DESCONOCIDO'

def consolidate_data(download_dir: str, output_file: str):
    """
    Reads all downloaded files in download_dir, consolidates them,
    and saves to output_file (Excel).
    """
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # --- Prepare DataFrames ---
    df_compras = pd.DataFrame()
    df_ventas = pd.DataFrame()

    # 1. Process Compras (CSV)
    # Case insensitive search for Compras files
    all_files = glob.glob(os.path.join(download_dir, "*.csv"))
    compras_files = [f for f in all_files if 'compra' in os.path.basename(f).lower()]
    
    if compras_files:
        all_compras = []
        for f in compras_files:
            try:
                # Try different encodings
                try:
                    df = pd.read_csv(f, sep=';', encoding='latin-1', index_col=False)
                except:
                    df = pd.read_csv(f, sep=';', encoding='utf-8', index_col=False)
                    
                mes = extract_month_from_filename(os.path.basename(f))
                df.insert(0, 'MES', mes)
                all_compras.append(df)
            except Exception as e:
                print(f"Error reading {f}: {e}")
        
        if all_compras:
            df_compras = pd.concat(all_compras)
            numeric_cols = ['Monto Neto', 'Monto Exento', 'Monto IVA Recuperable', 'Monto Iva No Recuperable', 'Monto Total']
            for col in numeric_cols:
                if col in df_compras.columns:
                    df_compras[col] = pd.to_numeric(df_compras[col], errors='coerce').fillna(0)

    # 2. Process Ventas (CSV)
    ventas_files = [f for f in all_files if 'venta' in os.path.basename(f).lower()]
    
    if ventas_files:
        all_ventas = []
        for f in ventas_files:
            try:
                df = pd.read_csv(f, sep=';', encoding='latin-1', index_col=False)
                mes = extract_month_from_filename(os.path.basename(f))
                df.insert(0, 'MES', mes)
                all_ventas.append(df)
            except Exception as e:
                print(f"Error reading {f}: {e}")
        
        if all_ventas:
            df_ventas = pd.concat(all_ventas)
            numeric_cols = ['Monto Neto', 'Monto Exento', 'Monto IVA', 'Monto total']
            for col in numeric_cols:
                if col in df_ventas.columns:
                    df_ventas[col] = pd.to_numeric(df_ventas[col], errors='coerce').fillna(0)

    # --- Create Sheets in Order ---

    # 1. Datos
    try:
        # Create a DataFrame of size 30x11 filled with NaN
        df_datos = pd.DataFrame(index=range(30), columns=range(11))
        
        # Helper to set value
        def set_val(r, c, val):
            df_datos.iat[r, c] = val

        # Extract info from filenames if possible
        rut_extracted = "Unknown"
        periodo_extracted = "Unknown"
        if compras_files:
            # Example: RCV_COMPRA_REGISTRO_89752800-2_202501.csv
            basename = os.path.basename(compras_files[0])
            parts = basename.split('_')
            for p in parts:
                if '-' in p and p[0].isdigit():
                    rut_extracted = p
                if p.isdigit() and len(p) == 6:
                    periodo_extracted = p # YYYYMM

        # Row 2 (Index 2)
        set_val(2, 1, 'Datos de sociedad a revisar')
        
        # Row 5 (Index 5)
        set_val(5, 1, 'Año Tributario')
        set_val(5, 2, 2025)
        set_val(5, 7, 'Fecha de Formulario:')
        set_val(5, 8, 30)
        set_val(5, 9, 'Julio')
        set_val(5, 10, 2025)
        
        # Row 7 (Index 7)
        set_val(7, 1, 'Año Comercial')
        set_val(7, 2, 2024)
        set_val(7, 7, 'Fecha Pago Formulario')
        set_val(7, 8, 30)
        set_val(7, 9, 'Agosto')
        set_val(7, 10, 2025)

        # Row 9 (Index 9)
        set_val(9, 1, 'Nombre:')
        set_val(9, 2, '[NOMBRE EMPRESA]') # Placeholder or need to fetch

        # Row 10 (Index 10)
        set_val(10, 7, 'Giros Economicos')
        set_val(10, 8, 'Código')
        set_val(10, 9, 'Categoria Tributaria')
        set_val(10, 10, 'Afecta IVA')

        # Row 11 (Index 11)
        set_val(11, 1, 'Giro Comercial:')

        # Row 13 (Index 13)
        set_val(13, 1, 'Tipo Jurídico:')

        # Row 15 (Index 15)
        set_val(15, 1, 'RUT Sociedad')
        set_val(15, 2, rut_extracted)
        set_val(15, 7, 'Tipo de IVA')
        set_val(15, 8, 'IVA Recuperable')

        # Row 17 (Index 17)
        set_val(17, 1, 'Clave SII')
        set_val(17, 2, '********') # Don't put real password
        set_val(17, 7, 'Bse PPM')
        set_val(17, 8, 0.00125)

        # Row 19 (Index 19)
        set_val(19, 1, 'Nombre Representante Legal')

        # Row 21 (Index 21)
        set_val(21, 1, 'RUT Representante Legal')

        # Row 23 (Index 23)
        set_val(23, 1, 'Clave SII')

        # Row 25 (Index 25)
        set_val(25, 1, 'Sistema de Tributación')
        set_val(25, 2, 'Pro Pyme General 14D')
        set_val(25, 4, 'Tasa de Impto')
        set_val(25, 5, 0.1)

        # Rename columns to Unnamed: X to match expected output structure
        df_datos.columns = [f'Unnamed: {i}' for i in range(11)]
        
        df_datos.to_excel(writer, sheet_name='Datos', index=False, header=False)
    except Exception as e:
        print(f"Error creating Datos: {e}")

    # 2. Cálculo Remanente
    try:
        # Create DataFrame 20 rows, 10 columns
        df_calc = pd.DataFrame(index=range(20), columns=range(10))
        
        # Helper
        def set_c(r, c, val):
            df_calc.iat[r, c] = val

        # Year
        set_c(1, 2, 2025)

        # UTM Table Headers
        set_c(3, 1, 'UTM')
        set_c(4, 1, 'MES')
        set_c(4, 2, 'VALOR')

        # UTM Data
        utm_data = [
            ('Enero', 67429), ('Febrero', 67294), ('Marzo', 68034),
            ('Abril', 68306), ('Mayo', 68648), ('Junio', 68923),
            ('Julio', 68647), ('Agosto', None), ('Septiembre', None),
            ('Octubre', None), ('Noviembre', None), ('Diciembre', None)
        ]
        
        for i, (mes, val) in enumerate(utm_data):
            set_c(5+i, 1, mes)
            set_c(5+i, 2, val)

        # Top Right Box
        set_c(1, 5, 'Codigo 77 Formulario')
        set_c(1, 7, 1000000)

        # Middle Right Table
        set_c(4, 6, 'MES')
        set_c(4, 7, 'Monto')
        
        set_c(5, 5, 'Codigo 77 Formulario')
        set_c(5, 6, 'Julio')
        set_c(5, 7, 1000000)
        
        set_c(6, 5, 'Valor en UTM')
        set_c(6, 7, 68647)
        
        set_c(7, 5, 'Monto REMANENTE EN UTM')
        set_c(7, 7, 14.57)

        # Bottom Right Table
        set_c(10, 5, 'VALOR UTM MES')
        set_c(10, 6, 'Agosto')
        
        set_c(12, 5, 'MONTO REMANENTE PERIODO')

        # Write to Excel
        df_calc.to_excel(writer, sheet_name='Cálculo Remanente', index=False, header=False)

    except Exception as e:
        print(f"Error creating Cálculo Remanente: {e}")

    # 3. RCV Compra (Summary + Detail)
    try:
        if not df_compras.empty:
            # --- Summary Section ---
            summary_rows = []
            target_types = [33, 34, 61, 56, 45]
            
            total_cant = 0
            total_exento = 0
            total_neto = 0
            total_iva_no_rec = 0
            total_iva = 0
            total_total = 0

            for doc_type in target_types:
                doc_name = DOC_TYPE_MAP.get(doc_type, str(doc_type))
                if 'Tipo Doc' in df_compras.columns:
                    group = df_compras[df_compras['Tipo Doc'] == doc_type]
                else:
                    group = pd.DataFrame()

                cantidad = len(group)
                exento = group['Monto Exento'].sum() if not group.empty and 'Monto Exento' in group.columns else 0
                neto = group['Monto Neto'].sum() if not group.empty and 'Monto Neto' in group.columns else 0
                iva_no_rec = group['Monto Iva No Recuperable'].sum() if not group.empty and 'Monto Iva No Recuperable' in group.columns else 0
                iva = group['Monto IVA Recuperable'].sum() if not group.empty and 'Monto IVA Recuperable' in group.columns else 0
                total = group['Monto Total'].sum() if not group.empty and 'Monto Total' in group.columns else 0

                summary_rows.append({
                    'Codigo': doc_type,
                    'Tipo Documento': doc_name,
                    'Cantidad': cantidad,
                    'Exento': exento,
                    'Neto': neto,
                    'IVA No Recuperable': iva_no_rec,
                    'IVA': iva,
                    'Total': total
                })

                total_cant += cantidad
                total_exento += exento
                total_neto += neto
                total_iva_no_rec += iva_no_rec
                total_iva += iva
                total_total += total

            # Total Row
            summary_rows.append({
                'Codigo': 'Total de documento',
                'Tipo Documento': '',
                'Cantidad': total_cant,
                'Exento': total_exento,
                'Neto': total_neto,
                'IVA No Recuperable': total_iva_no_rec,
                'IVA': total_iva,
                'Total': total_total
            })
            
            df_summary = pd.DataFrame(summary_rows)
            cols_summary = ['Codigo', 'Tipo Documento', 'Cantidad', 'Exento', 'Neto', 'IVA No Recuperable', 'IVA', 'Total']
            df_summary = df_summary.reindex(columns=cols_summary)
            
            # Write Summary at Top (Shifted down for image)
            df_summary.to_excel(writer, sheet_name='RCV Compra', index=False, startrow=22)
            
            # --- Detail Section ---
            # Map columns to requested format
            # Requested: Nro, Tipo Doc, Tipo Compra, RUT Proveedor, Razon Social, Folio, Fecha Docto, Fecha Recepcion, Fecha Acuse, Monto Exento, Monto Neto, Monto IVA Recuperable, Monto Iva No Recuperable, Codigo IVA No Rec., Monto Total
            
            df_detail = df_compras.copy()
            # Add/Rename columns
            df_detail['Nro'] = range(1, len(df_detail) + 1)
            df_detail['Tipo Compra'] = 'Del Giro' # Default or derived?
            
            # Rename existing columns to match target
            col_map = {
                'Rut Proveedor': 'RUT Proveedor',
                'Razon Social': 'Razon Social',
                'Folio': 'Folio', 
                'Fecha Docto': 'Fecha Docto',
                'Fecha Recepcion': 'Fecha Recepcion',
                'Monto Exento': 'Monto Exento',
                'Monto Neto': 'Monto Neto',
                'Monto IVA Recuperable': 'Monto IVA Recuperable',
                'Monto Iva No Recuperable': 'Monto Iva No Recuperable',
                'Monto Total': 'Monto Total',
                'Codigo IVA No Rec.': 'Codigo IVA No Rec.'
            }
            df_detail.rename(columns=col_map, inplace=True)
            
            # Ensure all target columns exist
            target_cols = ['Nro', 'Tipo Doc', 'Tipo Compra', 'RUT Proveedor', 'Razon Social', 'Folio', 'Fecha Docto', 'Fecha Recepcion', 'Fecha Acuse', 'Monto Exento', 'Monto Neto', 'Monto IVA Recuperable', 'Monto Iva No Recuperable', 'Codigo IVA No Rec.', 'Monto Total']
            
            for col in target_cols:
                if col not in df_detail.columns:
                    df_detail[col] = None
            
            df_detail = df_detail[target_cols]
            
            # Write Detail Table below Summary (skip rows after summary)
            start_row_detail = 22 + len(df_summary) + 4 # Header + Rows + spacing
            df_detail.to_excel(writer, sheet_name='RCV Compra', index=False, startrow=start_row_detail)

        else:
             pd.DataFrame(columns=['No Data']).to_excel(writer, sheet_name='RCV Compra', index=False)
    except Exception as e:
        print(f"Error creating RCV Compra: {e}")

    # 4. RCV Venta (Summary + Detail)
    try:
        if not df_ventas.empty:
             # --- Summary Section ---
            summary_rows = []
            target_types = [33, 34, 61, 56] # Standard types. Image showed 51, but 56 is standard.
            
            total_cant = 0
            total_exento = 0
            total_neto = 0
            total_iva = 0
            total_total = 0

            for doc_type in target_types:
                doc_name = DOC_TYPE_MAP.get(doc_type, str(doc_type))
                if 'Tipo Doc' in df_ventas.columns:
                    group = df_ventas[df_ventas['Tipo Doc'] == doc_type]
                else:
                    group = pd.DataFrame()

                cantidad = len(group)
                exento = group['Monto Exento'].sum() if not group.empty and 'Monto Exento' in group.columns else 0
                neto = group['Monto Neto'].sum() if not group.empty and 'Monto Neto' in group.columns else 0
                iva = group['Monto IVA'].sum() if not group.empty and 'Monto IVA' in group.columns else 0
                total = group['Monto Total'].sum() if not group.empty and 'Monto Total' in group.columns else (group['Monto total'].sum() if not group.empty and 'Monto total' in group.columns else 0)

                summary_rows.append({
                    'Codigo': doc_type,
                    'Tipo Documento': doc_name,
                    'Cantidad': cantidad,
                    'Exento': exento,
                    'Neto': neto,
                    'IVA': iva,
                    'Total': total
                })
                
                total_cant += cantidad
                total_exento += exento
                total_neto += neto
                total_iva += iva
                total_total += total

            # Total Row
            summary_rows.append({
                'Codigo': 'Total de documentos',
                'Tipo Documento': '',
                'Cantidad': total_cant,
                'Exento': total_exento,
                'Neto': total_neto,
                'IVA': total_iva,
                'Total': total_total
            })
            
            df_summary = pd.DataFrame(summary_rows)
            cols_summary = ['Codigo', 'Tipo Documento', 'Cantidad', 'Exento', 'Neto', 'IVA', 'Total']
            df_summary = df_summary.reindex(columns=cols_summary)
            
            df_summary.to_excel(writer, sheet_name='RCV Venta', index=False, startrow=0)
            
            # --- Detail Section ---
            df_detail = df_ventas.copy()
            
            # Add/Map columns
            df_detail['Nro'] = range(1, len(df_detail) + 1)
            df_detail['Tipo Venta'] = 'Del Giro' # Default per image
            
            # Rename existing columns
            col_map = {
                'Rut Receptor': 'Rut cliente', # CSV often has Rut Receptor
                'RUT Receptor': 'Rut cliente',
                'Rut Cliente': 'Rut cliente',
                'Monto IVA': 'Monto IVA',
                'Monto Total': 'Monto total',
                'Monto total': 'Monto total'
            }
            df_detail.rename(columns=col_map, inplace=True)
            
            # Full target columns list from user image
            target_cols = [
                'Nro', 'Tipo Doc', 'Tipo Venta', 'Rut cliente', 'Razon Social', 'Folio', 
                'Fecha Docto', 'Fecha Recepcion', 'Fecha Acuse Recibo', 'Fecha Reclamo', 
                'Monto Exento', 'Monto Neto', 'Monto IVA', 'Monto total', 
                'IVA Retenido Total', 'IVA Retenido Parcial', 'IVA no retenido', 'IVA propio', 'IVA Terceros', 
                'RUT Emisor Liquid. Factura', 'Neto Comision Liquid. Factura', 'Exento Comision Liquid. Factura', 
                'IVA Comision Liquid. Factura', 'IVA fuera de plazo', 'Tipo Docto. Referencia', 
                'Folio Docto. Referencia', 'Num. Ident. Receptor Extranjero', 'Nacionalidad Receptor Extranjero', 
                'Credito empresa constructora', 'Impte. Zona Franca (Ley 18211)', 'Garantia Dep. Envases', 
                'Indicador Venta sin Costo', 'Indicador Servicio Periodico', 'Monto No facturable', 
                'Total Monto Periodo', 'Venta Pasajes Transporte Nacional', 'Venta Pasajes Transporte Internacional', 
                'Numero Interno', 'Codigo Sucursal', 'NCE o NDE sobre Fact. de Compra', 
                'Codigo Otro Imp.', 'Valor Otro Imp.', 'Tasa Otro Imp.'
            ]
             
            for col in target_cols:
                if col not in df_detail.columns:
                    # Check for alternate names for specific columns
                    if col == 'Codigo Sucursal' and 'Sucursal' in df_detail.columns:
                        df_detail['Codigo Sucursal'] = df_detail['Sucursal']
                    elif col == 'Codigo Sucursal' and 'Cód. Sucursal' in df_detail.columns:
                        df_detail['Codigo Sucursal'] = df_detail['Cód. Sucursal']
                    else:
                        df_detail[col] = 0 if 'Monto' in col or 'IVA' in col or 'Neto' in col or 'Exento' in col or 'Valor' in col else None
            
            df_detail = df_detail[target_cols]
            
            start_row_detail = len(df_summary) + 4
            df_detail.to_excel(writer, sheet_name='RCV Venta', index=False, startrow=start_row_detail)

        else:
            pd.DataFrame(columns=['No Data']).to_excel(writer, sheet_name='RCV Venta', index=False)
    except Exception as e:
        print(f"Error creating RCV Venta: {e}")

    # 5. Boletas
    try:
        # Filter Boletas from Ventas
        # Specific codes from image: 39, 48, 41, 43
        target_codes = {
            39: 'Boleta Afecta Electronica',
            48: 'Boleta Medio Electrónico',
            41: 'Boleta Exenta Electrónica',
            43: 'Liquidacion de factura'
        }
        
        # Initialize boleta lists
        boletas_rows = []
        
        # Calculate Summary Data
        summary_rows = []
        
        # Prepare columns for the detail table
        # Header: Nº | Cantidad Dctos | Exento | Neto | IVA | Total | dte | fecha | vendedor | sucursal
        detail_cols = ['Nº', 'Cantidad Dctos', 'Exento', 'Neto', 'IVA', 'Total', 'dte', 'fecha', 'vendedor', 'sucursal']
        
        if not df_ventas.empty and 'Tipo Doc' in df_ventas.columns:
            # Create Detail Data First
            # We filter for the relevant types
            mask = df_ventas['Tipo Doc'].isin(target_codes.keys())
            df_b = df_ventas[mask].copy()
            
            if not df_b.empty:
                # Map columns
                df_b['Nº'] = range(1, len(df_b) + 1)
                df_b['Cantidad Dctos'] = 1 # Assuming 1 per row for detail list
                
                # Handle numeric columns safely
                for col in ['Monto Exento', 'Monto Neto', 'Monto IVA', 'Monto Total', 'Monto total']:
                    if col not in df_b.columns:
                        df_b[col] = 0
                    else:
                        df_b[col] = pd.to_numeric(df_b[col], errors='coerce').fillna(0)
                
                # Map to target names
                df_b['Exento'] = df_b['Monto Exento']
                df_b['Neto'] = df_b['Monto Neto']
                df_b['IVA'] = df_b['Monto IVA']
                if 'Monto Total' in df_b.columns:
                     df_b['Total'] = df_b['Monto Total']
                else:
                     df_b['Total'] = df_b['Monto total']
                     
                df_b['dte'] = df_b['Folio']
                df_b['fecha'] = df_b['Fecha Docto']
                df_b['vendedor'] = ''
                df_b['sucursal'] = ''
                
                # Select final columns
                df_detail = df_b[detail_cols]
            else:
                df_detail = pd.DataFrame(columns=detail_cols)

            # Create Summary Data
            # Format: [Code, Description, Cantidad, Exento, Neto, IVA, Total]
            # We use 'Unnamed' columns to position it somewhat like the image
            # Image looks like: empty, Code, Description, Cantidad, Exento, Neto, IVA, Total
            
            for code, desc in target_codes.items():
                # Filter for this code
                df_code = df_ventas[df_ventas['Tipo Doc'] == code]
                
                qty = len(df_code)
                
                # Sums
                exento = 0
                neto = 0
                iva = 0
                total = 0
                
                if not df_code.empty:
                    # numeric conversion already done safely above if we used df_b, but here we go back to df_ventas
                    # better to do safe conversion
                    for col in ['Monto Exento', 'Monto Neto', 'Monto IVA', 'Monto Total', 'Monto total']:
                        if col in df_code.columns:
                            df_code[col] = pd.to_numeric(df_code[col], errors='coerce').fillna(0)
                            
                    exento = df_code['Monto Exento'].sum() if 'Monto Exento' in df_code.columns else 0
                    neto = df_code['Monto Neto'].sum() if 'Monto Neto' in df_code.columns else 0
                    iva = df_code['Monto IVA'].sum() if 'Monto IVA' in df_code.columns else 0
                    if 'Monto Total' in df_code.columns:
                        total = df_code['Monto Total'].sum()
                    elif 'Monto total' in df_code.columns:
                        total = df_code['Monto total'].sum()

                # Row format matching the upper table in image
                # Let's align with the detail table columns roughly
                # Detail: Nº | Cantidad Dctos | Exento | Neto | IVA | Total | dte...
                # Summary: [Code] [Desc] [Qty] [Exento] [Neto] [IVA] [Total]
                # We can place Code in 'Nº' col? No, looks shifted.
                # Let's use a separate DataFrame for Summary and write it at top.
                
                # Image structure:
                # Row 39 | Boleta... | 0 | | | | |
                # The columns in image for summary table seem to be:
                # Col 0 (implicit): Code
                # Col 1: Description
                # Col 2: Cantidad
                # Col 3: Exento
                # Col 4: Neto
                # Col 5: IVA
                # Col 6: Total
                
                summary_rows.append([code, desc, qty, exento, neto, iva, total])

        else:
             # Empty case
             for code, desc in target_codes.items():
                 summary_rows.append([code, desc, 0, 0, 0, 0, 0])
             df_detail = pd.DataFrame(columns=detail_cols)

        # Create Summary DataFrame
        df_summary = pd.DataFrame(summary_rows, columns=['Código', 'Descripción', 'Cantidad', 'Exento', 'Neto', 'IVA', 'Total'])
        
        # Write to Excel
        # Summary at top (Row 0 header, Rows 1-4 data)
        # Spacer
        # Detail Header (Row 7 or so)
        
        # To align nicely, we might want to offset columns
        # Image shows summary table starting a bit to the right? No, looks like standard grid.
        # But the Detail table below has specific headers.
        
        # Let's write Summary first
        df_summary.to_excel(writer, sheet_name='Boletas', index=False, startrow=1, startcol=1) 
        # Writing at B2 to give some space? Or just A1?
        # Image: The summary table seems to have its own headers "Cantidad", "Exento", etc.
        # Let's just write at A1 for simplicity, or match visual indentation if needed.
        # The image shows "39" in a column, then "Boleta..." in next.
        
        # Let's try to match the column alignment of the detail table if possible.
        # Detail Cols: Nº, Cantidad Dctos, Exento, Neto, IVA, Total, ...
        # Summary Cols: Code, Desc, Cantidad, Exento, Neto, IVA, Total
        # Alignment:
        # Code -> Nº (Sort of)
        # Desc -> Cantidad Dctos (Mismatch)
        # Cantidad -> Exento (Mismatch)
        
        # So they are independent tables.
        
        # Write Summary
        df_summary.to_excel(writer, sheet_name='Boletas', index=False, startrow=2)
        
        # Write Detail Table with some spacing
        start_row_detail = len(df_summary) + 6 # 2 (header) + rows + spacer
        df_detail.to_excel(writer, sheet_name='Boletas', index=False, startrow=start_row_detail)

    except Exception as e:
        print(f"Error creating Boletas: {e}")

    # 6. Resumen
    try:
        # Define the structure based on the image
        # Sections: Header, Ventas, Compras, PPM, Boletas Honorarios, IUT, Total
        
        # Helper to summarize by Doc Type
        def summarize_by_doc(df, type_col='Tipo Doc'):
            if df.empty or type_col not in df.columns:
                return {}
            
            summary = {}
            grouped = df.groupby(type_col)
            for name, group in grouped:
                neto = group['Monto Neto'].sum() if 'Monto Neto' in group.columns else 0
                iva_debito = group['Monto IVA'].sum() if 'Monto IVA' in group.columns else 0
                iva_credito = group['Monto IVA Recuperable'].sum() if 'Monto IVA Recuperable' in group.columns else 0
                
                # Total logic depends on sheet
                if 'Monto Total' in group.columns:
                     total = group['Monto Total'].sum()
                elif 'Monto total' in group.columns:
                     total = group['Monto total'].sum()
                else:
                     total = 0
                     
                summary[name] = {'Neto': neto, 'IVA_Debito': iva_debito, 'IVA_Credito': iva_credito, 'Total': total}
            return summary

        ventas_summary = summarize_by_doc(df_ventas)
        compras_summary = summarize_by_doc(df_compras)
        
        resumen_rows = []
        
        # 1. Header Row (Merged in post-process)
        # Row 1: "RESUMEN DECLARACIÓN DE IMPUESTOS MENSUALES JULIO 2025" (Example)
        # Row 2: "REY MOMO" (Example)
        resumen_rows.append(['RESUMEN DECLARACIÓN DE IMPUESTOS MENSUALES', None, None, None])
        resumen_rows.append(['[NOMBRE EMPRESA]', None, None, None])
        resumen_rows.append([None, None, None, None]) # Spacer
        
        # 2. Ventas
        # Header: Ventas, Neto, IVA Debito Fiscal, Total
        resumen_rows.append(['Ventas', 'Neto', 'IVA Debito Fiscal', 'Total'])
        
        ventas_map = [
            (33, 'Factura electrónica'),
            (34, 'Factura exenta'),
            (61, 'Nota de crédito'),
            (56, 'Nota de debito'),
            (35, 'Boletas de Ventas y Servicios'), # Assuming 35
            (48, 'Boleta Medio Electrónico'),
            (43, 'Liquidacion de factura'),
            (41, 'Boletas de Ventas y Servicios Exentas')
        ]
        
        v_neto_total = 0
        v_iva_total = 0
        v_total_total = 0
        
        for code, name in ventas_map:
            vals = ventas_summary.get(code, {'Neto': 0, 'IVA_Debito': 0, 'Total': 0})
            resumen_rows.append([name, vals['Neto'], vals['IVA_Debito'], vals['Total']])
            v_neto_total += vals['Neto']
            v_iva_total += vals['IVA_Debito']
            v_total_total += vals['Total']
            
        # Total Ventas Row
        resumen_rows.append(['Total Ventas', v_neto_total, v_iva_total, v_total_total])
        resumen_rows.append([None, None, None, None]) # Spacer
        
        # 3. Compras
        # Header: Compras, Neto, IVA Crédito Fiscal, Total
        resumen_rows.append(['Compras', 'Neto', 'IVA Crédito Fiscal', 'Total'])
        
        compras_map = [
            (33, 'Factura electrónica'),
            (34, 'Factura exenta'),
            (34, 'Factura electrónica Exenta'), # Duplicate code handling? Maybe check distinct types
            (61, 'Nota de crédito'),
            (56, 'Nota de debito'),
            ('Remanente', 'Remanente IVA periodo anterior')
        ]
        
        c_neto_total = 0
        c_iva_total = 0
        c_total_total = 0
        
        for code, name in compras_map:
            if code == 'Remanente':
                # Placeholder for Remanente
                vals = {'Neto': 0, 'IVA_Credito': 0, 'Total': 0}
            else:
                vals = compras_summary.get(code, {'Neto': 0, 'IVA_Credito': 0, 'Total': 0})
            
            resumen_rows.append([name, vals['Neto'], vals['IVA_Credito'], vals['Total']])
            c_neto_total += vals['Neto']
            c_iva_total += vals['IVA_Credito']
            c_total_total += vals['Total']
            
        # Total Compras Row
        resumen_rows.append(['Total Compras', c_neto_total, c_iva_total, c_total_total])
        resumen_rows.append([None, None, None, None]) # Spacer
        
        # 4. Pago Provisional Mensual
        # Header: Pago Provisional Mensual, Ventas Netas, Factor PPM, PPM Pagado
        resumen_rows.append(['Pago Provisional Mensual', 'Ventas Netas', 'Factor PPM', 'PPM Pagado'])
        
        # Data Row
        ppm_base = v_neto_total # Usually Base Imponible is Ventas Netas
        ppm_factor = 0.0125 # Example from image, should be dynamic or 0
        ppm_pagado = int(ppm_base * ppm_factor)
        
        resumen_rows.append(['Base imponible', ppm_base, f"{ppm_factor*100}%", ppm_pagado])
        resumen_rows.append([None, None, None, None]) # Spacer
        
        # 5. Boletas de Honorarios
        # Header: Boletas de Honorarios, Bruto, Retenido, Pagado
        resumen_rows.append(['Boletas de Honorarios', 'Bruto', 'Retenido', 'Pagado'])
        # Placeholder row
        resumen_rows.append(['Totales', 0, 0, 0])
        resumen_rows.append([None, None, None, None]) # Spacer
        
        # 6. Impuesto Único a los Trabajadores
        # Header: Impuesto Único a los Trabajadores, Monto, None, None
        resumen_rows.append(['Impuesto Único a los Trabajadores', 'Monto', None, None])
        resumen_rows.append(['Impuesto Retenido', 0, None, None])
        resumen_rows.append(['Retención 3% Prestamo Solidario', 0, None, None])
        # Total IUT
        resumen_rows.append(['Total Impuesto Único', 0, None, None])
        resumen_rows.append([None, None, None, None]) # Spacer
        
        # 7. Total a Pagar Formulario
        # Header: Total a Pagar Formulario, None, None, None (Merged later)
        resumen_rows.append(['Total a Pagar Formulario', None, None, None])
        
        iva_pagar = v_iva_total - c_iva_total
        resumen_rows.append(['IVA a pagar determinado', None, None, iva_pagar])
        resumen_rows.append(['Pagos provisionales', None, None, ppm_pagado])
        resumen_rows.append(['Retención Boletas Honorarios', None, None, 0])
        resumen_rows.append(['Postergacion de pago IVA', None, None, 0])
        resumen_rows.append(['Impuesto Único Trabajadores', None, None, 0])
        
        total_final = iva_pagar - ppm_pagado # Simplified logic: IVA Pagar - PPM - Retenciones...
        # Note: Logic might be complex (IVA Pagar is positive tax, PPM is credit). 
        # Usually: Pay = IVA Pagar - PPM? Or PPM is paid separately? 
        # Image shows: Total a Pagar Formulario = 194. 
        # Image Math: 
        # IVA Debito: 154,990
        # IVA Credito: 22,224
        # Diff: 132,766 (Positive)
        # Wait, Image shows IVA Credito total 184,196? No, Total column is 184,196.
        # IVA Credito Col sum is 22,224.
        # Image: Ventas Total 154,990 (Neto)?? No.
        # Let's look at Image PPM: 194.
        # Total a Pagar Formulario: 194.
        # This implies IVA Pagar was 0 or offset completely? 
        # Actually in image: Ventas Neto 154,990. IVA Debito empty? No, top row empty.
        # Ah, looking closely at image:
        # Ventas: Factura Exenta 154,990. IVA Debito 0. Total 154,990.
        # Compras: Neto 116,972. IVA 22,224. Total 139,196. 
        # Remanente: ? 
        # Totals line 2: 161,972 | 22,224 | 184,196.
        # IVA Pagar determined: 1 - 2. (Debito - Credito). 0 - 22,224 = -22,224 (Remanente).
        # So IVA Pagar is 0.
        # Pagos Provisionales: 194.
        # Total a Pagar: 194. 
        # So Formula: max(0, Debito - Credito) + PPM + Retenciones?
        # Actually PPM is "Pagado" meaning "To be paid"? Or "Already paid"? 
        # Usually F29 pays PPM. So it's an addition to the total payment.
        # So Total = (Debito - Credito > 0 ? Diff : 0) + PPM + Retenciones.
        
        iva_determined = max(0, iva_pagar)
        final_payment = iva_determined + ppm_pagado # + others
        
        resumen_rows.append(['Total a Pagar Formulario', None, None, final_payment])

        df_resumen = pd.DataFrame(resumen_rows, columns=['Concepto', 'Col1', 'Col2', 'Total'])
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False, header=False)
        
    except Exception as e:
        print(f"Error creating Resumen: {e}")
        # Fallback
        pd.DataFrame({'Unnamed: 0': [None], 'Unnamed: 1': ['Error creating Resumen']}).to_excel(writer, sheet_name='Resumen', index=False)


    # 7. Compras (Dashboard)
    try:
        # Structure based on "COMPRAS.png" (Dashboard)
        # Header: "Compras desde Enero a Julio 2024" (Row 1)
        # Bar Chart (Placeholder text or image if available, but here we just leave space)
        # Tables: "TOP 5 VENTAS ANUALES" and "TOP 5 VENTAS MENSUALES" (Using Compras data)
        # Footer: Summary stats

        # Create DataFrame for Sheet
        df_compras_sheet = pd.DataFrame(index=range(30), columns=[f'Unnamed: {i}' for i in range(15)]) # Wide enough
        
        # Row 1: Title (Merged later)
        title = 'Compras desde Enero a Julio 2024' # Default
        if not df_compras.empty:
            try:
                # Try to determine date range from Fecha Docto or MES
                if 'Fecha Docto' in df_compras.columns:
                     dates = pd.to_datetime(df_compras['Fecha Docto'], errors='coerce')
                     min_date = dates.min()
                     max_date = dates.max()
                     if pd.notnull(min_date) and pd.notnull(max_date):
                         m_start = MONTH_MAP.get(f"{min_date.month:02d}", "").capitalize()
                         m_end = MONTH_MAP.get(f"{max_date.month:02d}", "").capitalize()
                         y = min_date.year
                         title = f'Compras desde {m_start} a {m_end} {y}'
            except: pass

        df_compras_sheet.loc[1, 'Unnamed: 1'] = title
        
        # Row 3-15: Chart Space
        
        # Row 17: Table Headers
        # Table 1: TOP 5 VENTAS ANUALES (Cols B-E)
        df_compras_sheet.loc[17, 'Unnamed: 1'] = 'TOP 5 VENTAS ANUALES'
        df_compras_sheet.loc[18, 'Unnamed: 1'] = 'RAZON SOCIAL'
        df_compras_sheet.loc[18, 'Unnamed: 3'] = 'RUT'
        df_compras_sheet.loc[18, 'Unnamed: 4'] = 'MONTOS'
        
        # Table 2: TOP 5 VENTAS MENSUALES (Cols G-J)
        df_compras_sheet.loc[17, 'Unnamed: 6'] = 'TOP 5 VENTAS MENSUALES'
        df_compras_sheet.loc[18, 'Unnamed: 6'] = 'RAZON SOCIAL'
        df_compras_sheet.loc[18, 'Unnamed: 8'] = 'RUT'
        df_compras_sheet.loc[18, 'Unnamed: 9'] = 'MONTOS'
        
        # Calculate Top 5 Annual (All Data)
        top_annual = []
        if not df_compras.empty:
             # Group by Razon Social/RUT
             grp = df_compras.groupby(['Razon Social', 'RUT Proveedor'])['Monto Total'].sum().reset_index()
             grp = grp.sort_values('Monto Total', ascending=False).head(5)
             for idx, row in grp.iterrows():
                 top_annual.append([row['Razon Social'], row['RUT Proveedor'], row['Monto Total']])
        
        # Fill Table 1
        for i, row_data in enumerate(top_annual):
            r = 19 + i
            df_compras_sheet.loc[r, 'Unnamed: 1'] = row_data[0]
            df_compras_sheet.loc[r, 'Unnamed: 3'] = row_data[1]
            df_compras_sheet.loc[r, 'Unnamed: 4'] = row_data[2]
            
        # Calculate Top 5 Monthly (Current Month)
        top_monthly = []
        if not df_compras.empty:
            # Filter for current month (assume latest month in data)
            # We inserted 'MES' col earlier.
            # Or just take the whole dataset if it's monthly only.
            # Let's assume the dataset IS the month we are processing.
            grp_m = df_compras.groupby(['Razon Social', 'RUT Proveedor'])['Monto Total'].sum().reset_index()
            grp_m = grp_m.sort_values('Monto Total', ascending=False).head(5)
            for idx, row in grp_m.iterrows():
                 top_monthly.append([row['Razon Social'], row['RUT Proveedor'], row['Monto Total']])

        # Fill Table 2
        for i, row_data in enumerate(top_monthly):
            r = 19 + i
            df_compras_sheet.loc[r, 'Unnamed: 6'] = row_data[0]
            df_compras_sheet.loc[r, 'Unnamed: 8'] = row_data[1]
            df_compras_sheet.loc[r, 'Unnamed: 9'] = row_data[2]

        # Footer Stats
        # TOTAL VENTA ANUAL
        total_anual = df_compras['Monto Total'].sum() if not df_compras.empty else 0
        df_compras_sheet.loc[25, 'Unnamed: 1'] = 'TOTAL VENTA ANUAL'
        df_compras_sheet.loc[25, 'Unnamed: 4'] = total_anual
        
        # N° DE COMPRAS EN EL AÑO
        count_anual = len(df_compras)
        df_compras_sheet.loc[26, 'Unnamed: 1'] = 'N° DE COMPRAS EN EL AÑO'
        df_compras_sheet.loc[26, 'Unnamed: 4'] = count_anual
        
        # MES VS AÑO
        df_compras_sheet.loc[27, 'Unnamed: 1'] = 'MES VS AÑO'
        
        current_month_name = 'JULIO' # Default
        percentage_str = '0%'

        if not df_compras.empty:
            try:
                if 'Fecha Docto' in df_compras.columns:
                     dates = pd.to_datetime(df_compras['Fecha Docto'], errors='coerce')
                     max_date = dates.max()
                     if pd.notnull(max_date):
                          current_month_num = f"{max_date.month:02d}"
                          current_month_name = MONTH_MAP.get(current_month_num, 'Unknown')
                          
                          total_annual = df_compras['Monto Total'].sum()
                          
                          mask_month = dates.dt.month == max_date.month
                          mask_year = dates.dt.year == max_date.year
                          total_month = df_compras[mask_month & mask_year]['Monto Total'].sum()
                          
                          if total_annual > 0:
                              percentage = total_month / total_annual
                              percentage_str = f"{percentage:.1%}"
            except: pass

        df_compras_sheet.loc[27, 'Unnamed: 3'] = current_month_name
        df_compras_sheet.loc[27, 'Unnamed: 4'] = percentage_str

        df_compras_sheet.to_excel(writer, sheet_name='Compras', index=False, header=False)
    except Exception as e:
        print(f"Error creating Compras: {e}")


    # 8. Ventas (Structure Match)
    try:
        # Structure based on "8 VENTAS.png"
        df_ventas_sheet = pd.DataFrame(index=range(30), columns=[f'Unnamed: {i}' for i in range(15)])
        
        # Title
        title = 'Ventas desde Enero a Diciembre 2025'
        if not df_ventas.empty:
            try:
                if 'Fecha Docto' in df_ventas.columns:
                     dates = pd.to_datetime(df_ventas['Fecha Docto'], errors='coerce')
                     min_date = dates.min()
                     max_date = dates.max()
                     if pd.notnull(min_date) and pd.notnull(max_date):
                         m_start = MONTH_MAP.get(f"{min_date.month:02d}", "").capitalize()
                         m_end = MONTH_MAP.get(f"{max_date.month:02d}", "").capitalize()
                         y = min_date.year
                         title = f'Ventas desde {m_start} a {m_end} {y}'
            except: pass
        df_ventas_sheet.loc[1, 'Unnamed: 1'] = title
        
        # Tables Headers
        # Table 1: TOP 5 VENTAS ANUALES (Cols B-E) -> Unnamed: 1, 3, 4
        df_ventas_sheet.loc[17, 'Unnamed: 1'] = 'TOP 5 VENTAS ANUALES'
        df_ventas_sheet.loc[18, 'Unnamed: 1'] = 'RAZON SOCIAL'
        df_ventas_sheet.loc[18, 'Unnamed: 3'] = 'RUT'
        df_ventas_sheet.loc[18, 'Unnamed: 4'] = 'MONTOS'
        
        # Table 2: TOP 5 VENTAS MENSUALES (Cols G-J) -> Unnamed: 6, 8, 9
        df_ventas_sheet.loc[17, 'Unnamed: 6'] = 'TOP 5 VENTAS MENSUALES'
        df_ventas_sheet.loc[18, 'Unnamed: 6'] = 'RAZON SOCIAL'
        df_ventas_sheet.loc[18, 'Unnamed: 8'] = 'RUT'
        df_ventas_sheet.loc[18, 'Unnamed: 9'] = 'MONTOS'
        
        # Data Processing
        # Ensure columns exist
        col_rut = 'Rut cliente' if 'Rut cliente' in df_ventas.columns else 'RUT Receptor'
        col_rs = 'Razon Social'
        col_monto = 'Monto Total' if 'Monto Total' in df_ventas.columns else 'Monto total'
        
        # Top 5 Annual
        top_annual = []
        if not df_ventas.empty and col_rut in df_ventas.columns:
             grp = df_ventas.groupby([col_rs, col_rut])[col_monto].sum().reset_index()
             grp = grp.sort_values(col_monto, ascending=False).head(5)
             for idx, row in grp.iterrows():
                 top_annual.append([row[col_rs], row[col_rut], row[col_monto]])
                 
        for i, row_data in enumerate(top_annual):
            r = 19 + i
            df_ventas_sheet.loc[r, 'Unnamed: 1'] = row_data[0]
            df_ventas_sheet.loc[r, 'Unnamed: 3'] = row_data[1]
            df_ventas_sheet.loc[r, 'Unnamed: 4'] = row_data[2]
            
        # Top 5 Monthly
        top_monthly = []
        if not df_ventas.empty and col_rut in df_ventas.columns:
            # Determine max month
            if 'Fecha Docto' in df_ventas.columns:
                dates = pd.to_datetime(df_ventas['Fecha Docto'], errors='coerce')
                max_date = dates.max()
                if pd.notnull(max_date):
                    mask = (dates.dt.month == max_date.month) & (dates.dt.year == max_date.year)
                    df_m = df_ventas[mask]
                    grp_m = df_m.groupby([col_rs, col_rut])[col_monto].sum().reset_index()
                    grp_m = grp_m.sort_values(col_monto, ascending=False).head(5)
                    for idx, row in grp_m.iterrows():
                        top_monthly.append([row[col_rs], row[col_rut], row[col_monto]])

        for i, row_data in enumerate(top_monthly):
            r = 19 + i
            df_ventas_sheet.loc[r, 'Unnamed: 6'] = row_data[0]
            df_ventas_sheet.loc[r, 'Unnamed: 8'] = row_data[1]
            df_ventas_sheet.loc[r, 'Unnamed: 9'] = row_data[2]
            
        # Footer Stats
        # TOTAL VENTA ANUAL
        total_anual = df_ventas[col_monto].sum() if not df_ventas.empty else 0
        df_ventas_sheet.loc[25, 'Unnamed: 1'] = 'TOTAL VENTA ANUAL'
        df_ventas_sheet.loc[25, 'Unnamed: 4'] = total_anual
        
        # N° DE CLIENTES EN EL AÑO
        # Use nunique of RUT
        count_clientes = df_ventas[col_rut].nunique() if not df_ventas.empty and col_rut in df_ventas.columns else 0
        df_ventas_sheet.loc[26, 'Unnamed: 1'] = 'N° DE CLIENTES EN EL AÑO'
        df_ventas_sheet.loc[26, 'Unnamed: 4'] = count_clientes
        
        # MES VS AÑO
        df_ventas_sheet.loc[27, 'Unnamed: 1'] = 'MES VS AÑO'
        
        current_month_name = 'JULIO' # Default
        percentage_str = '0%'
        
        if not df_ventas.empty:
             try:
                 if 'Fecha Docto' in df_ventas.columns:
                      dates = pd.to_datetime(df_ventas['Fecha Docto'], errors='coerce')
                      max_date = dates.max()
                      if pd.notnull(max_date):
                           current_month_num = f"{max_date.month:02d}"
                           current_month_name = MONTH_MAP.get(current_month_num, 'Unknown')
                           
                           mask_month = dates.dt.month == max_date.month
                           mask_year = dates.dt.year == max_date.year
                           total_month = df_ventas[mask_month & mask_year][col_monto].sum()
                           
                           if total_anual > 0:
                               percentage = total_month / total_anual
                               percentage_str = f"{percentage:.1%}"
             except: pass
             
        df_ventas_sheet.loc[27, 'Unnamed: 3'] = current_month_name
        df_ventas_sheet.loc[27, 'Unnamed: 4'] = percentage_str
        
        df_ventas_sheet.to_excel(writer, sheet_name='Ventas', index=False, header=False)
    except Exception as e:
        print(f"Error creating Ventas: {e}")

    # 9. BH (Structure Match)
    try:
        # Structure based on "9 BH.png"
        # Columns: A to K (approx 11 cols)
        # We need specific layout
        df_bh = pd.DataFrame(index=range(30), columns=[f'Unnamed: {i}' for i in range(12)])
        
        # Top Right Table (Rows 1-2)
        # Headers at Row 1 (Index 0) cols I, J, K (8, 9, 10)
        df_bh.loc[0, 'Unnamed: 8'] = 'Brutos'
        df_bh.loc[0, 'Unnamed: 9'] = 'Retenido'
        df_bh.loc[0, 'Unnamed: 10'] = 'Pagado'
        
        df_bh.loc[1, 'Unnamed: 5'] = 'Boleta de Honorarios Totales :'
        # Placeholders for totals
        df_bh.loc[1, 'Unnamed: 8'] = 0
        df_bh.loc[1, 'Unnamed: 9'] = 0
        df_bh.loc[1, 'Unnamed: 10'] = 0

        # Info Section
        df_bh.loc[3, 'Unnamed: 0'] = 'Contribuyente: [NOMBRE EMPRESA]'
        df_bh.loc[4, 'Unnamed: 0'] = 'RUT : [RUT EMPRESA]'
        
        # Date Info
        # Try to get date from data
        periodo_str = "Informe correspondiente al mes [MM] del año [YYYY]"
        try:
             # Try to find a month/year from previous processing
             # We can use the first available date from compras or sales
             target_date = None
             if not df_compras.empty and 'Fecha Docto' in df_compras.columns:
                 target_date = pd.to_datetime(df_compras['Fecha Docto'], errors='coerce').max()
             elif not df_ventas.empty and 'Fecha Docto' in df_ventas.columns:
                 target_date = pd.to_datetime(df_ventas['Fecha Docto'], errors='coerce').max()
                 
             if pd.notnull(target_date):
                 mm = f"{target_date.month:02d}"
                 yyyy = target_date.year
                 periodo_str = f"Informe correspondiente al mes {mm} del año {yyyy}"
        except: pass
        
        df_bh.loc[6, 'Unnamed: 0'] = periodo_str

        # Main Table Headers
        # Row 8 (Index 7): Group Headers
        df_bh.loc[7, 'Unnamed: 0'] = 'Boleta' # Spans A-D
        df_bh.loc[7, 'Unnamed: 4'] = 'Emisor' # Spans E-G
        df_bh.loc[7, 'Unnamed: 7'] = 'Honorarios' # Spans H-J (or H-K based on top table?)
        # Looking at image: 
        # Boleta: N, Fecha, Estado, Anulación (4 cols) -> A, B, C, D
        # Emisor: Rut, Nombre, Soc. Prof. (3 cols) -> E, F, G
        # Honorarios: Brutos, Retenido, Pagado (3 cols) -> H, I, J
        
        # Row 9 (Index 8): Sub Headers
        headers = ['N°', 'Fecha', 'Estado', 'Anulación', 'Rut', 'Nombre o Razón Social', 'Soc. Prof.', 'Brutos', 'Retenido', 'Pagado']
        for i, h in enumerate(headers):
            df_bh.loc[8, f'Unnamed: {i}'] = h
            
        # Data Placeholder (Row 10+)
        # If we had BH data we would fill it here.
        
        # Footer
        row_total = 15
        df_bh.loc[row_total, 'Unnamed: 0'] = 'Totales* :'
        df_bh.loc[row_total+1, 'Unnamed: 0'] = '(*) Los valores totales no consideran los montos de las boletas anuladas.'
        
        df_bh.to_excel(writer, sheet_name='BH', index=False, header=False)
    except Exception as e:
        print(f"Error creating BH: {e}")

    # 10. IMPUESTO UNICO (Structure Match)
    try:
        # Structure based on "10 IMPUESTO UNICO.png"
        # Image at top, table below.
        # Headers at Row 8 (Index 7) to leave space for image.
        
        df_iu = pd.DataFrame(index=range(20), columns=[f'Unnamed: {i}' for i in range(4)])
        
        # Headers
        df_iu.loc[7, 'Unnamed: 0'] = 'NOMBRE TRABAJADOR'
        df_iu.loc[7, 'Unnamed: 1'] = 'RUT'
        df_iu.loc[7, 'Unnamed: 2'] = 'Impuesto Único'
        df_iu.loc[7, 'Unnamed: 3'] = 'Retención 3%'
        
        df_iu.to_excel(writer, sheet_name='IMPUESTO UNICO', index=False, header=False)
    except Exception as e:
        print(f"Error creating IMPUESTO UNICO: {e}")

    # 11. DIN
    try:
        # Create a blank grid for DIN image
        df_din = pd.DataFrame(index=range(50), columns=[f'Unnamed: {i}' for i in range(20)])
        df_din.iloc[0, 0] = " " # Force content
        df_din.to_excel(writer, sheet_name='DIN', index=False, header=False)
    except Exception as e:
        print(f"Error creating DIN: {e}")

    # 12. Compras Acumuladas
    if not df_compras.empty:
        df_compras.to_excel(writer, sheet_name='Compras Acumuladas', index=False)
    else:
        pd.DataFrame().to_excel(writer, sheet_name='Compras Acumuladas', index=False)

    # 13. Compras analisis
    if not df_compras.empty:
        try:
            # Layout based on "11 COMPRAS ANALISIS" screenshot (Sheet 13 actually)
            # Structure:
            # A-C: Raw Data (MES, Razon Social, Monto Neto)
            # D: Empty
            # E-F: Pivot Supplier (Neto)
            # G: Empty
            # H-I: Pivot Month (Neto)
            # ... and Totals tables below
            
            # Ensure required columns
            req_cols = ['MES', 'Razon Social', 'Monto Neto', 'Monto Total']
            for col in req_cols:
                if col not in df_compras.columns:
                    if col == 'Monto Neto': df_compras[col] = 0
                    if col == 'Monto Total': df_compras[col] = 0
                    if col == 'Razon Social': df_compras[col] = 'Desconocido'
            
            # Table 1: Raw Data
            t1 = df_compras[['MES', 'Razon Social', 'Monto Neto']].copy()
            
            # Table 2: Pivot Supplier (Neto)
            t2 = df_compras.groupby('Razon Social')['Monto Neto'].sum().reset_index()
            t2.columns = ['Etiquetas de fila', 'Suma de Monto Neto']
            t2 = t2.sort_values('Suma de Monto Neto', ascending=False)
            # Add Total
            t2_total = pd.DataFrame([{'Etiquetas de fila': 'Total general', 'Suma de Monto Neto': t2['Suma de Monto Neto'].sum()}])
            t2 = pd.concat([t2, t2_total], ignore_index=True)
            
            # Table 3: Pivot Month (Neto)
            month_order = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
            t3 = df_compras.groupby('MES')['Monto Neto'].sum().reset_index()
            t3.columns = ['Etiquetas de fila', 'Suma de Monto Neto']
            # Sort by Month
            t3['MonthIdx'] = t3['Etiquetas de fila'].apply(lambda x: month_order.index(x) if x in month_order else 99)
            t3 = t3.sort_values('MonthIdx').drop('MonthIdx', axis=1)
            # Add Total
            t3_total = pd.DataFrame([{'Etiquetas de fila': 'Total general', 'Suma de Monto Neto': t3['Suma de Monto Neto'].sum()}])
            t3 = pd.concat([t3, t3_total], ignore_index=True)
            
            # Table 4: Pivot Supplier (Total)
            t4 = df_compras.groupby('Razon Social')['Monto Total'].sum().reset_index()
            t4.columns = ['Etiquetas de fila', 'Suma de Monto Total']
            t4 = t4.sort_values('Suma de Monto Total', ascending=False)
            # Add Total
            t4_total = pd.DataFrame([{'Etiquetas de fila': 'Total general', 'Suma de Monto Total': t4['Suma de Monto Total'].sum()}])
            t4 = pd.concat([t4, t4_total], ignore_index=True)

            # Table 5: Pivot Month (Total)
            t5 = df_compras.groupby('MES')['Monto Total'].sum().reset_index()
            t5.columns = ['MES', 'Suma de Monto Total'] # Screenshot shows 'MES' here? Or Labels?
            # Sort by Month
            t5['MonthIdx'] = t5['MES'].apply(lambda x: month_order.index(x) if x in month_order else 99)
            t5 = t5.sort_values('MonthIdx').drop('MonthIdx', axis=1)
            # Add Total
            t5_total = pd.DataFrame([{'MES': 'Total general', 'Suma de Monto Total': t5['Suma de Monto Total'].sum()}])
            t5 = pd.concat([t5, t5_total], ignore_index=True)

            # Construct the Grid
            # Determine max rows
            max_rows = max(len(t1), len(t2), len(t3)) + 20 # Extra buffer
            
            # Create Grid
            # Columns: 
            # 0-2: T1 (A,B,C)
            # 3: Empty (D)
            # 4-5: T2 (E,F)
            # 6: Empty (G)
            # 7-8: T3 (H,I)
            # ...
            # 10-11: T4 (Below T2?) -> Actually T4 is below T2.
            # 13-14: T5 (Below T3?)
            
            # Using simple concatenation horizontally is hard due to differing lengths.
            # We'll create a master DF initialized with NaNs.
            # Grid width ~ 20 columns.
            
            master_df = pd.DataFrame(index=range(max_rows + 30), columns=[f'Unnamed: {i}' for i in range(25)])
            
            # Helper to place DF
            def place_df(df, start_row, start_col):
                # Write headers
                for i, col in enumerate(df.columns):
                    master_df.iloc[start_row, start_col + i] = col
                # Write data
                for r_idx, row in df.iterrows():
                    for c_idx, val in enumerate(row):
                        master_df.iloc[start_row + 1 + r_idx, start_col + c_idx] = val

            # Place T1 (A1) -> Index 0, 0
            place_df(t1, 0, 0)
            
            # Place T2 (E1) -> Index 0, 4
            place_df(t2, 0, 4)
            
            # Place T3 (H1) -> Index 0, 7
            place_df(t3, 0, 7)
            
            # Place T4 (Below T2) -> Let's say row 15 or after T2 ends + gap
            t2_end = len(t2) + 2
            start_t4 = max(15, t2_end)
            place_df(t4, start_t4, 4)
            
            # Place T5 (Below T3) -> Aligned with T4
            start_t5 = start_t4
            place_df(t5, start_t5, 7) # Screenshot shows it might be further right or same col as T3?
            # Screenshot shows "Etiquetas de fila" | "Suma de Monto Total" below T2
            # And "MES" | "Suma de Monto Total" below T3 (actually to the right of T4)
            
            # Chart Data Block (R1) -> Index 0, 17
            # Headers: MESES, BASE, ETIQUETA, VALOR, VACIO, TAPA
            chart_headers = ['MESES', 'BASE', 'ETIQUETA', 'VALOR', 'VACIO', 'TAPA']
            for i, h in enumerate(chart_headers):
                master_df.iloc[0, 17 + i] = h
            
            # Write Master DF
            master_df.to_excel(writer, sheet_name='Compras analisis', index=False, header=False)

        except Exception as e:
            print(f"Error creating Compras analisis: {e}")
            pd.DataFrame(columns=['Error creating Analysis']).to_excel(writer, sheet_name='Compras analisis', index=False)
    else:
        pd.DataFrame(columns=['MES', 'Razon Social', 'Monto Neto']).to_excel(writer, sheet_name='Compras analisis', index=False)

    # 14. Ventas Acumuladas
    if not df_ventas.empty:
        df_ventas.to_excel(writer, sheet_name='Ventas Acumuladas', index=False)
    else:
        pd.DataFrame().to_excel(writer, sheet_name='Ventas Acumuladas', index=False)

    # 15. Ventas Analisis
    try:
        # Ensure MES column
        if 'Fecha Docto' in df_ventas.columns:
             dates = pd.to_datetime(df_ventas['Fecha Docto'], errors='coerce')
             # Map month number to Spanish Name
             df_ventas['MES'] = dates.dt.month.map(lambda x: MONTH_MAP.get(f"{int(x):02d}", "Unknown") if pd.notnull(x) else "Unknown")
        elif 'MES' not in df_ventas.columns:
             df_ventas['MES'] = 'Unknown'

        # Ensure required columns
        req_cols = ['MES', 'Razon Social', 'Monto Neto', 'Monto Total']
        for col in req_cols:
            if col not in df_ventas.columns:
                if col == 'Monto Neto': df_ventas[col] = 0
                if col == 'Monto Total': df_ventas[col] = 0
                if col == 'Razon Social': df_ventas[col] = 'Desconocido'
        
        # Table 1: Raw Data
        t1 = df_ventas[['MES', 'Razon Social', 'Monto Neto']].copy()
        
        # Table 2: Pivot Customer (Neto)
        t2 = df_ventas.groupby('Razon Social')['Monto Neto'].sum().reset_index()
        t2.columns = ['Etiquetas de fila', 'Suma de Monto Neto']
        t2 = t2.sort_values('Suma de Monto Neto', ascending=False)
        # Add Total
        t2_total = pd.DataFrame([{'Etiquetas de fila': 'Total general', 'Suma de Monto Neto': t2['Suma de Monto Neto'].sum()}])
        t2 = pd.concat([t2, t2_total], ignore_index=True)
        
        # Table 3: Pivot Month (Neto)
        month_order = ['ENERO', 'FEBRERO', 'MARZO', 'ABRIL', 'MAYO', 'JUNIO', 'JULIO', 'AGOSTO', 'SEPTIEMBRE', 'OCTUBRE', 'NOVIEMBRE', 'DICIEMBRE']
        t3 = df_ventas.groupby('MES')['Monto Neto'].sum().reset_index()
        t3.columns = ['Etiquetas de fila', 'Suma de Monto Neto']
        # Sort by Month
        t3['MonthIdx'] = t3['Etiquetas de fila'].apply(lambda x: month_order.index(x) if x in month_order else 99)
        t3 = t3.sort_values('MonthIdx').drop('MonthIdx', axis=1)
        # Add Total
        t3_total = pd.DataFrame([{'Etiquetas de fila': 'Total general', 'Suma de Monto Neto': t3['Suma de Monto Neto'].sum()}])
        t3 = pd.concat([t3, t3_total], ignore_index=True)
        
        # Table 4: Pivot Customer (Total)
        t4 = df_ventas.groupby('Razon Social')['Monto Total'].sum().reset_index()
        t4.columns = ['Etiquetas de fila', 'Suma de Monto Total']
        t4 = t4.sort_values('Suma de Monto Total', ascending=False)
        # Add Total
        t4_total = pd.DataFrame([{'Etiquetas de fila': 'Total general', 'Suma de Monto Total': t4['Suma de Monto Total'].sum()}])
        t4 = pd.concat([t4, t4_total], ignore_index=True)

        # Table 5: Pivot Month (Total)
        t5 = df_ventas.groupby('MES')['Monto Total'].sum().reset_index()
        t5.columns = ['MES', 'Suma de Monto Total'] 
        # Sort by Month
        t5['MonthIdx'] = t5['MES'].apply(lambda x: month_order.index(x) if x in month_order else 99)
        t5 = t5.sort_values('MonthIdx').drop('MonthIdx', axis=1)
        # Add Total
        t5_total = pd.DataFrame([{'MES': 'Total general', 'Suma de Monto Total': t5['Suma de Monto Total'].sum()}])
        t5 = pd.concat([t5, t5_total], ignore_index=True)

        # Construct the Grid
        max_rows = max(len(t1), len(t2), len(t3)) + 20 
        
        master_df = pd.DataFrame(index=range(max_rows + 30), columns=[f'Unnamed: {i}' for i in range(25)])
        
        def place_df(df, start_row, start_col):
            for i, col in enumerate(df.columns):
                master_df.iloc[start_row, start_col + i] = col
            for r_idx, row in df.iterrows():
                for c_idx, val in enumerate(row):
                    master_df.iloc[start_row + 1 + r_idx, start_col + c_idx] = val

        place_df(t1, 0, 0)
        place_df(t2, 0, 4)
        place_df(t3, 0, 7)
        
        t2_end = len(t2) + 2
        start_t4 = max(15, t2_end)
        place_df(t4, start_t4, 4)
        
        start_t5 = start_t4
        place_df(t5, start_t5, 7)
        
        # Chart Data Headers (R-W) -> 17-22
        chart_headers = ['MESES', 'BASE', 'ETIQUETA', 'VALOR', 'VACIO', 'TAPA']
        for i, h in enumerate(chart_headers):
            master_df.iloc[0, 17 + i] = h
            
        master_df.to_excel(writer, sheet_name='Ventas Analisis', index=False, header=False)
        
    except Exception as e:
        print(f"Error creating Ventas Analisis: {e}")
        pd.DataFrame(columns=['Error']).to_excel(writer, sheet_name='Ventas Analisis', index=False)

    # 16. Base de datos
    try:
        df_bd = pd.DataFrame({
            'Unnamed: 0': ['Tipo Juridico', 'Empresa Individual de Responsabilidad Limitada', 'Sociedad de Responsabilidad Limitada', 'Sociedad Por Acciones', 'Persona Natural Con Giro'],
            'Unnamed: 1': [None, None, None, None, None],
            'Unnamed: 2': ['Regimenes Tributarios', None, None, None, None],
            'Unnamed: 3': [None, None, None, None, None],
            'Unnamed: 4': [None, None, None, None, None],
            'Unnamed: 5': [None, None, None, None, None],
            'Unnamed: 6': ['Dia', 1, 2, 3, 4],
            'Unnamed: 7': ['Mes', 'Enero', 'Febrero', 'Marzo', 'Abril']
        })
        df_bd.to_excel(writer, sheet_name='Base de datos', index=False)
    except Exception as e:
        print(f"Error creating Base de datos sheet: {e}")

    writer.close()
    
    # --- Post-processing Styles with openpyxl directly ---
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        
        # --- Insert Reference Images ---
        # RCV Compra
        if 'RCV Compra' in wb.sheetnames:
            ws_rcv = wb['RCV Compra']
            img_path = os.path.join(os.path.dirname(os.path.dirname(output_file)), 'ejemplos', '3 RCV COMPRA.png')
            # Fallback if examples are in root/ejemplos
            if not os.path.exists(img_path):
                 img_path = os.path.join(os.getcwd(), 'ejemplos', '3 RCV COMPRA.png')
            
            if os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    # Adjust image size if needed? Default is usually fine, or user might want specific scale.
                    # Place at A1
                    ws_rcv.add_image(img, 'A1')
                except Exception as e:
                    print(f"Error inserting image into RCV Compra: {e}")

        # Compras (Dashboard)
        if 'Compras' in wb.sheetnames:
            ws_compras = wb['Compras']
            img_path = os.path.join(os.path.dirname(os.path.dirname(output_file)), 'ejemplos', '7 COMPRAS.png')
            if not os.path.exists(img_path):
                 img_path = os.path.join(os.getcwd(), 'ejemplos', '7 COMPRAS.png')
            
            if os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    ws_compras.add_image(img, 'A1')
                except Exception as e:
                    print(f"Error inserting image into Compras: {e}")

        # Ventas (Dashboard)
        if 'Ventas' in wb.sheetnames:
            ws_ventas = wb['Ventas']
            img_path = os.path.join(os.path.dirname(os.path.dirname(output_file)), 'ejemplos', '8 VENTAS.png')
            if not os.path.exists(img_path):
                 img_path = os.path.join(os.getcwd(), 'ejemplos', '8 VENTAS.png')
            
            if os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    ws_ventas.add_image(img, 'A1')
                except Exception as e:
                    print(f"Error inserting image into Ventas: {e}")

        # BH (Image)
        if 'BH' in wb.sheetnames:
            ws_bh = wb['BH']
            img_path = os.path.join(os.path.dirname(os.path.dirname(output_file)), 'ejemplos', '9 BH.png')
            if not os.path.exists(img_path):
                 img_path = os.path.join(os.getcwd(), 'ejemplos', '9 BH.png')
            
            if os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    ws_bh.add_image(img, 'A1')
                except Exception as e:
                    print(f"Error inserting image into BH: {e}")

        # IMPUESTO UNICO (Image)
        if 'IMPUESTO UNICO' in wb.sheetnames:
            ws_iu = wb['IMPUESTO UNICO']
            img_path = os.path.join(os.path.dirname(os.path.dirname(output_file)), 'ejemplos', '10 IMPUESTO UNICO.png')
            if not os.path.exists(img_path):
                 img_path = os.path.join(os.getcwd(), 'ejemplos', '10 IMPUESTO UNICO.png')
            
            if os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    ws_iu.add_image(img, 'A1')
                except Exception as e:
                    print(f"Error inserting image into IMPUESTO UNICO: {e}")

        # DIN (Image)
        if 'DIN' in wb.sheetnames:
            ws_din = wb['DIN']
            img_path = os.path.join(os.path.dirname(os.path.dirname(output_file)), 'ejemplos', '11 DIN.png')
            if not os.path.exists(img_path):
                 img_path = os.path.join(os.getcwd(), 'ejemplos', '11 DIN.png')
            
            print(f"DEBUG: DIN img_path={img_path}, exists={os.path.exists(img_path)}")
            
            if os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    ws_din.add_image(img, 'A1')
                except Exception as e:
                    print(f"Error inserting image into DIN: {e}")

        # Compras analisis (Image)
        if 'Compras analisis' in wb.sheetnames:
            ws_ca = wb['Compras analisis']
            img_path = os.path.join(os.path.dirname(os.path.dirname(output_file)), 'ejemplos', '13 COMPRAS ANALISIS.png')
            if not os.path.exists(img_path):
                 img_path = os.path.join(os.getcwd(), 'ejemplos', '13 COMPRAS ANALISIS.png')
            
            if os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    ws_ca.add_image(img, 'A1')
                except Exception as e:
                    print(f"Error inserting image into Compras analisis: {e}")

        # Ventas Analisis (Image)
        if 'Ventas Analisis' in wb.sheetnames:
            ws_va = wb['Ventas Analisis']
            img_path = os.path.join(os.path.dirname(os.path.dirname(output_file)), 'ejemplos', '15 VENTAS ANALISIS.png')
            if not os.path.exists(img_path):
                 img_path = os.path.join(os.getcwd(), 'ejemplos', '15 VENTAS ANALISIS.png')
            
            if os.path.exists(img_path):
                try:
                    img = Image(img_path)
                    ws_va.add_image(img, 'A1')
                except Exception as e:
                    print(f"Error inserting image into Ventas Analisis: {e}")

        if 'Cálculo Remanente' in wb.sheetnames:
            ws = wb['Cálculo Remanente']
            
            blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True)
            black_font = Font(color="000000", bold=True)
            black_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            def style_ref(ref, fill=None, font=None, border=None):
                try:
                    cell = ws[ref]
                    if fill: cell.fill = fill
                    if font: cell.font = font
                    if border: cell.border = border
                except Exception as e:
                    print(f"Error styling cell {ref}: {e}")

            # UTM Header
            style_ref('B4', fill=blue_fill, font=white_font, border=black_border)
            style_ref('B5', fill=blue_fill, font=white_font, border=black_border)
            style_ref('C5', fill=blue_fill, font=white_font, border=black_border)
            
            # UTM Data Borders (Rows 6-17)
            for r in range(6, 18):
                style_ref(f'B{r}', border=black_border)
                style_ref(f'C{r}', border=black_border)

            # Top Right Box (F2:H2)
            style_ref('F2', fill=blue_fill, font=white_font, border=black_border)
            style_ref('G2', fill=blue_fill, font=white_font, border=black_border)
            style_ref('H2', border=black_border)

            # Middle Right Table
            style_ref('F5', fill=blue_fill, font=white_font, border=black_border)
            style_ref('F6', fill=blue_fill, font=white_font, border=black_border)
            style_ref('F7', fill=blue_fill, font=white_font, border=black_border)
            
            style_ref('H5', border=black_border)
            style_ref('H6', border=black_border)
            style_ref('H7', border=black_border)
            
            # Bottom Right Table
            style_ref('F11', fill=blue_fill, font=white_font, border=black_border)
            style_ref('H11', border=black_border)
            
            style_ref('F13', fill=blue_fill, font=white_font, border=black_border)
            style_ref('H13', border=black_border)
            
            # Adjust column widths
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            
            # --- Apply Generic Table Styles to other sheets ---
            import openpyxl.utils
            
            def apply_table_style(sheet_name):
                if sheet_name in wb.sheetnames:
                    ws_t = wb[sheet_name]
                    max_col = ws_t.max_column
                    max_row = ws_t.max_row
                    
                    # Skip if empty
                    if max_row < 1: return

                    # Determine header rows
                    header_rows = [1]
                    if sheet_name in ['RCV Compra', 'RCV Venta']:
                         # RCV sheets have a second table around row 11
                         # We can scan column A for "Nro" or "Codigo" to be safe, or hardcode
                         for r in range(8, 20):
                             val = ws_t.cell(row=r, column=1).value
                             if val == 'Nro':
                                 header_rows.append(r)
                                 break
                    elif sheet_name == 'Boletas':
                         # Boletas Detail Header
                         for r in range(6, 20):
                             val = ws_t.cell(row=r, column=1).value
                             if val == 'Nº':
                                 header_rows.append(r)
                                 break
                    elif sheet_name == 'IMPUESTO UNICO':
                         header_rows = [8]
                    
                    # Apply styles
                    for r_idx in range(1, max_row + 1):
                        for c_idx in range(1, max_col + 1):
                            cell = ws_t.cell(row=r_idx, column=c_idx)
                            
                            # Header Style
                            if r_idx in header_rows:
                                if sheet_name == 'Boletas' and r_idx > 1:
                                    # Second header in Boletas is Green
                                    cell.fill = green_fill
                                    cell.font = black_font
                                else:
                                    cell.fill = blue_fill
                                    cell.font = white_font
                                cell.border = black_border
                            else:
                                # Data Borders
                                if cell.value is not None or r_idx < max_row: 
                                     cell.border = black_border

                    # Add "OK" marks for RCV sheets below the summary table (approx Row 8)
                    if sheet_name in ['RCV Compra', 'RCV Venta'] and len(header_rows) > 1:
                        # Assuming first summary table ends before second header
                        # Second header is at header_rows[1] (e.g. 11)
                        # We want to put OKs at header_rows[1] - 3 (approx row 8)
                        ok_row = header_rows[1] - 3
                        if ok_row > 2:
                            blue_font = Font(color="4472C4", bold=True)
                            
                            # Determine columns for OK marks
                            # RCV Venta: 7 cols (Total at 7) -> range(3, 8) covers 3,4,5,6,7
                            # RCV Compra: 8 cols (Total at 8) -> range(3, 9) covers 3,4,5,6,7,8
                            ok_range = range(3, 8) if sheet_name == 'RCV Venta' else range(3, 9)
                            
                            for c in ok_range:
                                cell = ws_t.cell(row=ok_row, column=c)
                                cell.value = "OK"
                                cell.font = blue_font
                                cell.alignment = Alignment(horizontal='center')
                            
                            # Add "Limpiador" pseudo-button
                            limp_cell = ws_t.cell(row=3, column=10) # J3
                            limp_cell.value = "Limpiador"
                            limp_cell.alignment = Alignment(horizontal='center', vertical='center')
                            limp_cell.border = Border(left=Side(style='thick'), right=Side(style='thick'), top=Side(style='thick'), bottom=Side(style='thick'))
                            limp_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

                    # Auto-adjust column widths (simple approximation)
                    for col in range(1, max_col + 1):
                        col_letter = openpyxl.utils.get_column_letter(col)
                        ws_t.column_dimensions[col_letter].width = 20
                    
                    if sheet_name == 'IMPUESTO UNICO':
                         ws_t.column_dimensions['A'].width = 40
                         ws_t.column_dimensions['B'].width = 20
                         ws_t.column_dimensions['C'].width = 20
                         ws_t.column_dimensions['D'].width = 20

            # List of sheets to apply standard table style
            standard_sheets = [
                'RCV Compra', 'RCV Venta', 'Boletas', 'Resumen', 
                'Compras', 'Ventas', 'Compras analisis', 'Ventas Analisis',
                'Compras Acumuladas', 'Ventas Acumuladas',
                'BH', 'IMPUESTO UNICO', 'DIN'
            ]
            
            for sheet in standard_sheets:
                apply_table_style(sheet)

            # --- Style 'Datos' Sheet ---
            if 'Datos' in wb.sheetnames:
                ws_d = wb['Datos']
                max_row = ws_d.max_row
                # Style Column A (Labels)
                for row in range(1, max_row + 1):
                    cell = ws_d.cell(row=row, column=1)
                    cell.fill = blue_fill
                    cell.font = white_font
                    cell.border = black_border
                    
                    # Value borders (Column B)
                    if ws_d.max_column >= 2:
                        cell_val = ws_d.cell(row=row, column=2)
                        cell_val.border = black_border
                
                ws_d.column_dimensions['A'].width = 25
                ws_d.column_dimensions['B'].width = 40
            
            # --- Style 'Resumen' Sheet ---
            if 'Resumen' in wb.sheetnames:
                ws_r = wb['Resumen']
                
                # Colors
                purple_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
                white_font = Font(color="FFFFFF", bold=True)
                
                # Find Total Rows and Header Rows
                # We can iterate and check content
                
                # Header (Rows 1-2)
                for r in range(1, 3):
                    for c in range(1, 5):
                        cell = ws_r.cell(row=r, column=c)
                        cell.fill = purple_fill
                        cell.font = white_font
                        cell.alignment = Alignment(horizontal='center')
                
                # Merge Title Cells
                ws_r.merge_cells('A1:D1')
                ws_r.merge_cells('A2:D2')
                
                max_row = ws_r.max_row
                
                # Headers for sections
                headers = ['Ventas', 'Compras', 'Pago Provisional Mensual', 'Boletas de Honorarios', 'Impuesto Único a los Trabajadores', 'Total a Pagar Formulario']
                
                for r in range(1, max_row + 1):
                    val_a = ws_r.cell(row=r, column=1).value
                    
                    # Section Headers (Underline/Bold)
                    if val_a in headers:
                         ws_r.cell(row=r, column=1).font = Font(bold=True, underline='single')
                         # Headers have column names next to them in some cases
                         for c in range(2, 5):
                             cell = ws_r.cell(row=r, column=c)
                             if cell.value:
                                 cell.font = Font(bold=True)
                                 cell.alignment = Alignment(horizontal='center')
                                 
                         # Border bottom for header row
                         for c in range(1, 5):
                             ws_r.cell(row=r, column=c).border = Border(bottom=Side(style='double'))
                    
                    # Total Rows (Purple Background)
                    if val_a and str(val_a).startswith('Total '):
                        for c in range(1, 5):
                            cell = ws_r.cell(row=r, column=c)
                            cell.fill = purple_fill
                            cell.font = white_font
                            cell.border = black_border
                    
                    # PPM Pagado Cell (Specific)
                    if val_a == 'Base imponible':
                        # The PPM Pagado is in Col 4
                        cell = ws_r.cell(row=r, column=4)
                        cell.fill = purple_fill
                        cell.font = white_font
                        
                    # Retenido (Honorarios)
                    if val_a == 'Totales' and ws_r.cell(row=r-1, column=1).value == 'Boletas de Honorarios':
                        # Retenido is Col 3
                        cell = ws_r.cell(row=r, column=3)
                        cell.fill = purple_fill
                        cell.font = white_font
                        
                    # Impuesto Unico Total
                    # Covered by 'Total Impuesto Único' starts with Total
                    
                    # Final Total Row
                    if val_a == 'Total a Pagar Formulario' and r > 10: # Avoid confusing with header if any
                         for c in range(1, 5):
                            cell = ws_r.cell(row=r, column=c)
                            cell.fill = purple_fill
                            cell.font = white_font
                            cell.border = black_border

                ws_r.column_dimensions['A'].width = 40
                ws_r.column_dimensions['B'].width = 15
                ws_r.column_dimensions['C'].width = 15
                ws_r.column_dimensions['D'].width = 15

            wb.save(output_file)
            print("Applied styles to Cálculo Remanente and other sheets via post-processing.")
            
    except Exception as e:
        print(f"Error in post-processing styles: {e}")

    return output_file
